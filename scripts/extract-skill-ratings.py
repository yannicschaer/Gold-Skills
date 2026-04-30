#!/usr/bin/env python3
"""
Extracts skill ratings from the Excel skill matrix and writes a JSON file
that the Node migration script can consume.

Excel layout (per Ist/Soll sheet):
  row 1: category names (sparse, only on first column of each block)
  row 2: skill titles (cols C..AR)
  row 3: column headers ("Name", "Rolle")
  row 4..N: person rows — col A = name, col B = role, cols C..AR = ratings 0–5

Usage:
  python3 scripts/extract-skill-ratings.py
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl missing. Install with: pip3 install openpyxl")

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "gold-skills-imports" / "Design Team Skillmatrix.xlsx"
OUT = REPO / "gold-skills-imports" / "skill-ratings.json"

IST_SHEET = "Skillmatrix Ist 2025"
SOLL_SHEET = "Skillmatrix Soll 2026"
TEAM_SHEET = "Team"

FIRST_DATA_ROW = 4  # 1-indexed: rows 1+2 = header, row 3 = "Name/Rolle"
FIRST_SKILL_COL = 3  # 1-indexed col C
LAST_SKILL_COL = 44  # 1-indexed col AR

STOP_MARKERS = ("Schnitt", "Cluster", "=")


def read_skill_columns(ws):
    """Return list of (category, title) per column index 0..N covering FIRST_SKILL_COL..LAST_SKILL_COL."""
    row1 = [c.value for c in ws[1]]  # categories (sparse)
    row2 = [c.value for c in ws[2]]  # skill titles

    skills = []
    current_cat = None
    for col_idx in range(FIRST_SKILL_COL - 1, LAST_SKILL_COL):
        cat = row1[col_idx] if col_idx < len(row1) else None
        if cat is not None and str(cat).strip():
            current_cat = str(cat).strip()
        title = row2[col_idx] if col_idx < len(row2) else None
        if title is None or not str(title).strip():
            skills.append(None)
        else:
            skills.append({"category": current_cat, "title": str(title).strip()})
    return skills


def read_person_rows(ws):
    """Yield (name, role, [values]) for each person row until a stop marker."""
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        name = row[0]
        if name is None:
            # Empty row — skip but keep going (next row might be data); however
            # in practice the empty row precedes the stop markers, so we stop.
            return
        if any(m in str(name) for m in STOP_MARKERS):
            return
        role = row[1]
        values = list(row[FIRST_SKILL_COL - 1 : LAST_SKILL_COL])
        yield str(name).strip(), (str(role).strip() if role else None), values


def main():
    if not XLSX.exists():
        sys.exit(f"Excel file not found: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # 1) People + emails from "Team" sheet (col A = name, col B = email)
    name_to_email = {}
    team_ws = wb[TEAM_SHEET]
    for row in team_ws.iter_rows(min_row=2, values_only=True):
        name, email = row[0], row[1]
        if name and email:
            name_to_email[str(name).strip()] = str(email).strip().lower()

    # 2) Skills (use Ist as canonical — Soll has identical layout)
    ist_ws = wb[IST_SHEET]
    soll_ws = wb[SOLL_SHEET]
    skills = read_skill_columns(ist_ws)
    soll_skills = read_skill_columns(soll_ws)
    if [s["title"] if s else None for s in skills] != [
        s["title"] if s else None for s in soll_skills
    ]:
        sys.exit("Skill columns differ between Ist and Soll sheets — aborting.")

    # 3) Per-person ratings
    ist_by_name = {name: vals for name, _, vals in read_person_rows(ist_ws)}
    soll_by_name = {name: vals for name, _, vals in read_person_rows(soll_ws)}
    roles_by_name = {name: role for name, role, _ in read_person_rows(ist_ws)}
    for name, role, _ in read_person_rows(soll_ws):
        roles_by_name.setdefault(name, role)

    all_names = sorted(set(ist_by_name) | set(soll_by_name))

    people = []
    ratings_out = {}
    skipped_no_data = []

    for name in all_names:
        ist_vals = ist_by_name.get(name) or [None] * len(skills)
        soll_vals = soll_by_name.get(name) or [None] * len(skills)

        # Skip if neither sheet has any value for this person
        has_ist = any(v is not None for v in ist_vals)
        has_soll = any(v is not None for v in soll_vals)
        if not has_ist and not has_soll:
            skipped_no_data.append(name)
            continue

        per_skill = {}
        for idx, skill in enumerate(skills):
            if skill is None:
                continue
            cur = ist_vals[idx]
            tgt = soll_vals[idx]
            if cur is None and tgt is None:
                continue
            per_skill[skill["title"]] = {
                "current": int(cur) if cur is not None else None,
                "target": int(tgt) if tgt is not None else None,
            }

        people.append(
            {
                "name": name,
                "role": roles_by_name.get(name),
                "email": name_to_email.get(name),
            }
        )
        ratings_out[name] = per_skill

    payload = {
        "skills": [s for s in skills if s is not None],
        "people": people,
        "ratings": ratings_out,
    }

    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2))

    print(f"✅ Wrote {OUT}")
    print(f"   {len(payload['skills'])} skills")
    print(f"   {len(payload['people'])} people with ratings:")
    for p in payload["people"]:
        email = p["email"] or "<no email in Team sheet>"
        print(f"     - {p['name']:<25} {email}")
    if skipped_no_data:
        print(f"   skipped (no ratings): {', '.join(skipped_no_data)}")


if __name__ == "__main__":
    main()
